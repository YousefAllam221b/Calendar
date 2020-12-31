from __future__ import print_function
import datetime
import pickle
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

# If modifying these scopes, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/calendar.readonly']

#Prints 10 upcoming events
def main():
    """Shows basic usage of the Google Calendar API.
    Prints the start and name of the next 10 events on the user's calendar.
    """
    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    service = build('calendar', 'v3', credentials=creds)

    # Call the Calendar API
    now = datetime.datetime.utcnow().isoformat() + 'Z' # 'Z' indicates UTC time
    print('Getting the upcoming 10 events')
    events_result = service.events().list(calendarId='primary', timeMin=now,
                                        maxResults=10, singleEvents=True,
                                        orderBy='startTime').execute()
    events = events_result.get('items', [])

    if not events:
        print('No upcoming events found.')
    for event in events:
        start = event['start'].get('dateTime', event['start'].get('date'))
        print(start, event['summary'])

if __name__ == '__main__':
    main()

r'''
import openpyxl
from openpyxl import Workbook
book = Workbook()
sheet = book.active
fname = r'C:\Users\Yousef\Documents\CalendarApp\Schedule.xlsx'
wb = openpyxl.load_workbook(fname)


sheet1 = wb['Sheet1']
slots = [("1st",'08:15:00','09:45:00'),('2nd','10:00:00','11:30:00'),('3rd','11:45:00','13:15:00'),('4th','13:45:00','15:15:00'),('5th','15:45:00','17:15:00')]
#gettingSlots creates returns a list with tuples in form of (tut name , room , start , end)
def gettingSlots():
    alist = []
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if sheet1.cell(row=x,column =y).value != None:
                text = sheet1.cell(row=x,column =y).value.split(' ')
                start = ''
                end = ''
                for slot in slots:
                    if(str(slot[0]) == str(sheet1.cell(row=sheet1.min_row,column=y).value)):
                        start= slot[1]
                        end = slot[2]
                a = (text[0]+' ' +text[1],text[3],start,end)
                alist.append(a)
    return alist
'''
