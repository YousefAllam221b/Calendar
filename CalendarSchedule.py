from __future__ import print_function
import httplib2
import os

from apiclient import discovery
import oauth2client
from oauth2client import client
from oauth2client import tools
from oauth2client import file
import datetime

try:
    import argparse
    flags = argparse.ArgumentParser(parents=[tools.argparser]).parse_args()
except ImportError:
    flags = None

# If modifying these scopes, delete your previously saved credentials
# at ~/.credentials/calendar-python-quickstart.json
SCOPES = 'https://www.googleapis.com/auth/calendar'
CLIENT_SECRET_FILE = 'client_secret.json'
APPLICATION_NAME = 'Google Calendar API Python Quickstart'


def get_credentials():
    """Gets valid user credentials from storage.

    If nothing has been stored, or if the stored credentials are invalid,
    the OAuth2 flow is completed to obtain the new credentials.

    Returns:
        Credentials, the obtained credential.
    """
    home_dir = os.path.expanduser('~')
    credential_dir = os.path.join(home_dir, '.credentials')
    if not os.path.exists(credential_dir):
        os.makedirs(credential_dir)
    credential_path = os.path.join(credential_dir,
                                   'calendar-python-quickstart.json')

    store = oauth2client.file.Storage(credential_path)
    credentials = store.get()
    if not credentials or credentials.invalid:
        flow = client.flow_from_clientsecrets(CLIENT_SECRET_FILE, SCOPES)
        flow.user_agent = APPLICATION_NAME
        if flags:
            credentials = tools.run_flow(flow, store, flags)
        else: # Needed only for compatibility with Python 2.6
            credentials = tools.run(flow, store)
        print('Storing credentials to ' + credential_path)
    return credentials

import openpyxl
from openpyxl import Workbook
book = Workbook()
sheet = book.active
fname = r'C:\Users\Yousef\Documents\CalendarApp\Schedule.xlsx'
wb = openpyxl.load_workbook(fname)
sheet1 = wb['Sheet1']
days=[()]
slots = [("1st",'08:15:00','09:45:00'),('2nd','10:00:00','11:30:00'),('3rd','11:45:00','13:15:00'),('4th','13:45:00','15:15:00'),('5th','15:45:00','17:15:00')]
#gettingSlots creates returns a list with tuples in form of (tut name , room , start , end)
def gettingSlots():
    alist = []
    dict ={}
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        dict[sheet1.cell(row=x,column =sheet1.min_column).value]= []

        alist= []
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if sheet1.cell(row=x,column =y).value != None:
                text = sheet1.cell(row=x,column =y).value.split(' ')
                start = ''
                end = ''
                for slot in slots:
                    if(str(slot[0]) == str(sheet1.cell(row=sheet1.min_row,column=y).value)):
                        start= slot[1]
                        end = slot[2]
                a = (text[0]+' ' +text[1],text[3],start,end)
                alist.append(a)
        dict[sheet1.cell(row=x,column =sheet1.min_column).value] = alist
    return dict

weekdays=["saturday","sunday","monday","tuesday","wednesday","thursday"]
def Weekdays(weekdays):
    theWeekdays=[]
    holidays=[]
    flag=True
    for x in range(sheet1.min_row+1,sheet1.max_row+1):
        for y in range(sheet1.min_column+1,sheet1.max_column+1):
            if sheet1.cell(row=x,column =y).value != None:
                flag = False
        if flag==True:
            holidays+=[sheet1.cell(row=x,column =sheet1.min_column).value]
            flag=False
        else:
            theWeekdays+=[sheet1.cell(row=x,column =sheet1.min_column).value]
        flag=True
    return theWeekdays,holidays #returns weekdays while the other returns holidays


def main():
    """Shows basic usage of the Google Calendar API.

    Creates a Google Calendar API service object and outputs a list of the next
    10 events on the user's calendar.
    """
    credentials = get_credentials()
    http = credentials.authorize(httplib2.Http())
    service = discovery.build('calendar', 'v3', http=http)

    # Refer to the Python quickstart on how to setup the environment:
    # https://developers.google.com/google-apps/calendar/quickstart/python
    # Change the scope to 'https://www.googleapis.com/auth/calendar' and delete any
    # stored credentials.
    slots = gettingSlots()
    daysValues = {"Saturday": 0,"Sunday":1,"Monday":2,"Tuesday":3,"Wednesday":4,"Thursday":5}
    for key , value in slots.items():
        for slot in value:
            hms = slot[2].split(":")
            hms2 = slot[2].split(":")

            dateS = str(datetime.datetime(2021,1,2,int(hms[0]),int(hms[1]),int(hms[2]))+datetime.timedelta(days=daysValues[key])).split(' ')
            dateE = str(datetime.datetime(2021,1,2,int(hms2[0]),int(hms2[1]),int(hms2[2]))+datetime.timedelta(days=daysValues[key])).split(' ')
            event = {
              'summary': slot[0],
              'location': slot[1],
              'description': 'Testing app',
              'start': {
                'dateTime': dateS[0] + 'T' + dateS[1] + '+02:00',
                'timeZone': 'Africa/Cairo',
              },
              'end': {
                'dateTime': dateE[0]+'T'+ dateE[1] +'+02:00',
                'timeZone': 'Africa/Cairo',
              },
              'recurrence': [
                'RRULE:FREQ=WEEKLY;COUNT=12'
              ],
              'reminders': {
                'useDefault': False,
                'overrides': [
                  {'method': 'popup', 'minutes': 10},
                ],
              },
            }
            event = service.events().insert(calendarId='primary', body=event).execute()
            print ('Event created: %s' % (event.get('htmlLink')))
if __name__ == '__main__':
    main()
